"""Extract, clean, audit, and publish category taxonomy labels from Oracle.

This stage is designed to run before build_taxonomy_reference.py.

Usage:
    uv run extract_taxonomy_source.py --category BEER

Published workbook:
    data/categories/<CATEGORY>/<source.taxonomy_file>

Invalid placeholder labels are governed by one shared constant in this script;
category TOML files do not need taxonomy_extract or invalid_labels settings.

Run-scoped audit files:
    data/categories/<CATEGORY>/runs/<run_id>/taxonomy/
        taxonomy_source_raw.parquet
        taxonomy_source_cleaned.parquet
        taxonomy_source_rejected.parquet
        taxonomy_source_rejected.xlsx
        taxonomy_extract_summary.json
"""
from __future__ import annotations

import argparse
import json
import math
import re
import sys
import tomllib
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TRENDING_ROOT = Path(__file__).resolve().parent
WORKSPACE_ROOT = TRENDING_ROOT.parent

for required in (
    WORKSPACE_ROOT / "src" / "connection.py",
    TRENDING_ROOT / "core" / "project_context.py",
    TRENDING_ROOT / "core" / "run_manifest.py",
):
    if not required.exists():
        raise FileNotFoundError(f"缺少依赖文件：{required}")

for root in (WORKSPACE_ROOT, TRENDING_ROOT):
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))

from src.connection import create_oracle_connection_pool, execute_query  # noqa: E402
from core.project_context import ProjectContext  # noqa: E402
from core.run_manifest import create_manifest, update_stage  # noqa: E402


INVALID_LABELS = {
    "不知道",
    "没有注明",
    "未注明",
    "没注明",
    "未知",
    "不详",
    "不适用",
    "其它",
    "其他",
    "多种",
    "unknown",
    "other",
    "n/a",
    "na",
    "null",
}

SQL = """
SELECT DISTINCT
    t2.SEGTYPE AS SEGTYPE,
    t1.segdesc AS SEGDESC,
    t2.csegment AS CSEGMENT
FROM db_cate_segment t1
JOIN db_dic_segment t2
  ON t1.segno = t2.segno
 AND t1.catcode = t2.catcode
WHERE t1.catcode = :category_code
ORDER BY t2.SEGTYPE, t2.csegment
"""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract and clean one category taxonomy from Oracle"
    )
    parser.add_argument("--category", default=None, help="品类代码，例如 BEER")
    parser.add_argument("--run-id", default=None, help="可选：指定或覆盖 run_id")
    parser.add_argument(
        "--service",
        default=None,
        choices=("02", "03"),
        help="可选：覆盖 TOML 中的 Oracle service",
    )
    parser.add_argument(
        "--category-code",
        default=None,
        help="可选：覆盖 SQL 使用的 taxonomy category code",
    )
    return parser.parse_args()


def load_context(args: argparse.Namespace) -> ProjectContext:
    context = (
        ProjectContext.from_category(args.category, project_root=TRENDING_ROOT)
        if args.category
        else ProjectContext.active(project_root=TRENDING_ROOT)
    )
    return context.with_run_id(args.run_id) if args.run_id else context


def normalized_text(value: Any) -> str:
    """Normalize Oracle/Excel scalars to stable text without changing meaning."""
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    if isinstance(value, Decimal):
        text = format(value, "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def comparison_key(value: Any) -> str:
    return normalized_text(value).casefold()


def clean_taxonomy(
    raw: pd.DataFrame,
    *,
    excluded_dimension_codes: set[str],
    invalid_labels: set[str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    expected = ["SEGTYPE", "SEGDESC", "CSEGMENT"]
    missing = [column for column in expected if column not in raw.columns]
    if missing:
        raise KeyError(f"Oracle 结果缺少字段：{missing}")

    frame = raw[expected].copy()
    for column in expected:
        frame[column] = frame[column].map(normalized_text)

    invalid_keys = {comparison_key(value) for value in invalid_labels}
    excluded_codes = {comparison_key(value) for value in excluded_dimension_codes}

    accepted_rows: list[dict[str, str]] = []
    rejected_rows: list[dict[str, str]] = []

    for row_number, row in enumerate(frame.to_dict("records"), start=1):
        code = row["SEGTYPE"]
        name = row["SEGDESC"]
        label = row["CSEGMENT"]
        reasons: list[str] = []

        if not code:
            reasons.append("EMPTY_ATTRIBUTE_CODE")
        if not name:
            reasons.append("EMPTY_ATTRIBUTE_NAME")
        if not label:
            reasons.append("EMPTY_LABEL")
        if comparison_key(code) in excluded_codes:
            reasons.append("EXCLUDED_DIMENSION")
        if comparison_key(label) in invalid_keys:
            reasons.append("INVALID_PLACEHOLDER_LABEL")
        if label and comparison_key(label) == comparison_key(code):
            reasons.append("LABEL_EQUALS_ATTRIBUTE_CODE")
        if label and comparison_key(label) == comparison_key(name):
            reasons.append("LABEL_EQUALS_ATTRIBUTE_NAME")

        output = {
            "SEGTYPE": code,
            "SEGDESC": name,
            "CSEGMENT": label,
        }
        if reasons:
            rejected_rows.append(
                {
                    "SOURCE_ROW_NUMBER": row_number,
                    **output,
                    "REJECTION_REASON": ";".join(reasons),
                }
            )
        else:
            accepted_rows.append(output)

    cleaned = pd.DataFrame(accepted_rows, columns=expected)
    rejected = pd.DataFrame(
        rejected_rows,
        columns=["SOURCE_ROW_NUMBER", *expected, "REJECTION_REASON"],
    )

    if not cleaned.empty:
        cleaned = (
            cleaned.drop_duplicates(subset=expected, keep="first")
            .sort_values(["SEGTYPE", "CSEGMENT"], kind="stable")
            .reset_index(drop=True)
        )
    return cleaned, rejected


def extract_category_name(raw: pd.DataFrame) -> str:
    """Read the unique Chinese category name from SEGTYPE=CATEGORY."""
    missing = {"SEGTYPE", "CSEGMENT"} - set(raw.columns)
    if missing:
        raise KeyError(f"Oracle 结果缺少品类名称字段：{sorted(missing)}")
    names = {
        normalized_text(value)
        for value in raw.loc[
            raw["SEGTYPE"].map(comparison_key).eq("category"), "CSEGMENT"
        ].tolist()
        if normalized_text(value)
    }
    if not names:
        raise RuntimeError("未找到 SEGTYPE=CATEGORY 的 CSEGMENT，无法自动填写品类中文名。")
    if len(names) != 1:
        raise RuntimeError(f"SEGTYPE=CATEGORY 对应多个 CSEGMENT：{sorted(names)}")
    return next(iter(names))


def _replace_or_insert_toml_value(text: str, *, section: str, key: str, value: str) -> str:
    lines = text.splitlines()
    header = f"[{section}]"
    encoded = json.dumps(value, ensure_ascii=False)
    start = next((i for i, line in enumerate(lines) if line.strip() == header), None)
    if start is None:
        if lines and lines[-1].strip(): lines.append("")
        lines.extend([header, f"{key} = {encoded}"])
        return "\n".join(lines) + "\n"
    end = next((i for i in range(start + 1, len(lines)) if lines[i].strip().startswith("[")), len(lines))
    pattern = re.compile(rf"^\s*{re.escape(key)}\s*=")
    for i in range(start + 1, end):
        if pattern.match(lines[i]):
            lines[i] = f"{key} = {encoded}"
            return "\n".join(lines) + "\n"
    lines.insert(start + 1, f"{key} = {encoded}")
    return "\n".join(lines) + "\n"


def publish_category_metadata(*, category_code: str, category_name: str, config_path: Path, registry_path: Path) -> None:
    """Synchronize Oracle category metadata to category TOML and registry."""
    if not config_path.exists():
        raise FileNotFoundError(f"未找到品类配置：{config_path}")
    config_text = config_path.read_text(encoding="utf-8-sig")
    updated_config = _replace_or_insert_toml_value(config_text, section="category", key="name", value=category_name)
    tomllib.loads(updated_config)
    config_path.write_text(updated_config, encoding="utf-8")

    registry_text = registry_path.read_text(encoding="utf-8-sig") if registry_path.exists() else "[defaults]\nactive_category = " + json.dumps(category_code, ensure_ascii=False) + "\n"
    section = f"categories.{category_code}"
    is_new = f"[{section}]" not in registry_text
    updated_registry = _replace_or_insert_toml_value(registry_text, section=section, key="name", value=category_name)
    updated_registry = _replace_or_insert_toml_value(updated_registry, section=section, key="config", value=f"configs/{category_code}.toml")
    if is_new:
        updated_registry = _replace_or_insert_toml_value(updated_registry, section=section, key="enabled", value="true").replace('enabled = "true"','enabled = true')
    tomllib.loads(updated_registry)
    registry_path.parent.mkdir(parents=True, exist_ok=True)
    registry_path.write_text(updated_registry, encoding="utf-8")


def write_workbook(path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    imported_font = Font(color="008000")

    for column_index, column_name in enumerate(frame.columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = imported_font
            cell.alignment = Alignment(vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, column_name in enumerate(frame.columns, start=1):
        values = [str(column_name)] + [
            normalized_text(value)
            for value in frame.iloc[:, column_index - 1].head(2000).tolist()
        ]
        width = min(max(len(value) for value in values) + 2, 45)
        sheet.column_dimensions[get_column_letter(column_index)].width = max(width, 12)
    sheet.row_dimensions[1].height = 24
    workbook.save(path)


def main() -> None:
    args = parse_args()
    context = load_context(args)
    context.ensure_directories()
    category_config_path = TRENDING_ROOT / "configs" / f"{context.category_code}.toml"
    category_registry_path = TRENDING_ROOT / "configs" / "category_registry.toml"
    create_manifest(context)

    taxonomy_config = dict(context.config.get("taxonomy", {}))
    source_config = dict(context.config.get("source", {}))
    etl_config = dict(context.config.get("etl", {}))

    service = str(
        args.service
        or etl_config.get("service", "03")
    )
    category_code = str(
        args.category_code
        or source_config.get("taxonomy_catcode")
        or source_config.get("bundle")
        or context.category_code
    )
    excluded_dimensions = set(
        taxonomy_config.get(
            "excluded_dimension_codes",
            ["CATEGORY", "BRAND", "SUBBRAND"],
        )
    )
    # One shared cleaning contract for every category. Invalid placeholder
    # values are deliberately defined in this script, not in category TOML.
    invalid_labels = set(INVALID_LABELS)

    raw_path = context.taxonomy_dir / "taxonomy_source_raw.parquet"
    cleaned_path = context.taxonomy_dir / "taxonomy_source_cleaned.parquet"
    rejected_path = context.taxonomy_dir / "taxonomy_source_rejected.parquet"
    rejected_xlsx = context.taxonomy_dir / "taxonomy_source_rejected.xlsx"
    summary_path = context.taxonomy_dir / "taxonomy_extract_summary.json"
    published_xlsx = context.taxonomy_source_file

    update_stage(context, "taxonomy", "extracting_source")
    pool = None
    try:
        pool = create_oracle_connection_pool(service=service)
        rows, header = execute_query(
            pool,
            SQL,
            params={"category_code": category_code},
            return_header=True,
        )
        raw = pd.DataFrame(rows, columns=[str(value).upper() for value in header])
        category_name = extract_category_name(raw)
        publish_category_metadata(category_code=context.category_code, category_name=category_name, config_path=category_config_path, registry_path=category_registry_path)
        raw.to_parquet(raw_path, index=False)

        cleaned, rejected = clean_taxonomy(
            raw,
            excluded_dimension_codes=excluded_dimensions,
            invalid_labels=invalid_labels,
        )
        if cleaned.empty:
            raise RuntimeError("清洗后没有任何有效 Taxonomy 标签，拒绝覆盖正式文件。")

        cleaned.to_parquet(cleaned_path, index=False)
        rejected.to_parquet(rejected_path, index=False)
        write_workbook(published_xlsx, "Sheet1", cleaned)
        write_workbook(rejected_xlsx, "Rejected", rejected)

        reason_counts: dict[str, int] = {}
        for reason_text in rejected.get("REJECTION_REASON", pd.Series(dtype=str)).tolist():
            for reason in str(reason_text).split(";"):
                if reason:
                    reason_counts[reason] = reason_counts.get(reason, 0) + 1

        summary = {
            "status": "success",
            "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "category_code": context.category_code,
            "category_name": category_name,
            "category_config": str(category_config_path),
            "category_registry": str(category_registry_path),
            "taxonomy_category_code": category_code,
            "run_id": context.run_id,
            "oracle_service": service,
            "raw_rows": int(len(raw)),
            "cleaned_rows": int(len(cleaned)),
            "rejected_rows": int(len(rejected)),
            "rejected_by_reason": reason_counts,
            "excluded_dimension_codes": sorted(excluded_dimensions),
            "invalid_labels_exact_match": sorted(invalid_labels),
            "published_taxonomy_file": str(published_xlsx),
            "raw_parquet": str(raw_path),
            "cleaned_parquet": str(cleaned_path),
            "rejected_parquet": str(rejected_path),
            "rejected_workbook": str(rejected_xlsx),
            "note": "无效值使用标准化后的精确匹配；不会因包含‘其它’或‘无’而误删复合业务标签。",
        }
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        update_stage(
            context,
            "taxonomy",
            "source_extracted",
            artifacts={
                "taxonomy_source_published": published_xlsx,
                "taxonomy_source_raw": raw_path,
                "taxonomy_source_cleaned": cleaned_path,
                "taxonomy_source_rejected": rejected_path,
                "taxonomy_source_rejected_xlsx": rejected_xlsx,
                "taxonomy_extract_summary": summary_path,
            },
        )
        print(
            f"[OK] category={context.category_code} raw={len(raw)} "
            f"cleaned={len(cleaned)} rejected={len(rejected)}"
        )
        print(f"Published taxonomy: {published_xlsx}")
        print(f"Rejected audit: {rejected_xlsx}")
        print(
            "Next: uv run build_taxonomy_reference.py "
            f"--category {context.category_code}"
        )
    except Exception:
        update_stage(context, "taxonomy", "failed")
        raise
    finally:
        if pool is not None:
            try:
                pool.close()
            except Exception:
                pass


if __name__ == "__main__":
    main()
